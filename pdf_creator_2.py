# =======================================================importing libraries=============================================

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, protection, Font
import os


def layout(invoice_details):
    # ========================================================initialising workbook==========================================

    wb = openpyxl.load_workbook('Experiment.xlsx')
    ws = wb.active
    print("Worksheet created")

    customer_details = invoice_details['Customer_details']

    # =========================================================styles========================================================

    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 2
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 19
    ws.row_dimensions[6].height = 19
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 16
    ws.row_dimensions[9].height = 16
    ws.row_dimensions[10].height = 16
    ws.row_dimensions[11].height = 16
    ws.row_dimensions[12].height = 29
    for i in range(13, 23):
        ws.row_dimensions[i].height = 28

    ws.row_dimensions[23].height = 15
    for i in range(24, 28):
        ws.row_dimensions[i].height = 16

    ws.row_dimensions[28].height = 24
    for i in range(29, 33):
        ws.row_dimensions[i].height = 15

    ws.row_dimensions[34].height = 42
    ws.row_dimensions[35].height = 16
    ws.row_dimensions[36].height = 16
    ws.row_dimensions[37].height = 22
    ws.row_dimensions[38].height = 19
    ws.row_dimensions[39].height = 19
    ws.row_dimensions[40].height = 16
    ws.row_dimensions[41].height = 16
    ws.row_dimensions[42].height = 16
    ws.row_dimensions[43].height = 16
    ws.row_dimensions[44].height = 16
    ws.row_dimensions[45].height = 29
    for i in range(46, 56):
        ws.row_dimensions[i].height = 28

    ws.row_dimensions[56].height = 15
    for i in range(57, 61):
        ws.row_dimensions[i].height = 16

    ws.row_dimensions[61].height = 24
    for i in range(62, 66):
        ws.row_dimensions[i].height = 15

    thin = Side(border_style="thin", color="365194")
    table_border = Side(border_style="thin", color="B2BEB5")
    thick = Side(border_style="thick", color="365194")

    data_heading4 = Font(name='Constantia',
                         size=18,
                         color="365194")
    data_heading3 = Font(name='Franklin Gothic Book',
                         size=10)
    data_heading2 = Font(name='Franklin Gothic Book',
                         size=12)
    data_heading1 = Font(name='Constantia',
                         size=28)
    table_headings = Font(name='Constantia',
                          size=12,
                          bold=True,
                          color="FFFFFF")
    billing_data_headings = Font(name='Franklin Gothic Book',
                                 size=11,
                                 color="365194")
    billing_data = Font(name='Franklin Gothic Book',
                        size=11)

    ws['B1'].font = data_heading1
    ws['B1'].alignment = Alignment(horizontal='left', vertical='bottom')

    ws['B2'].font = data_heading2
    ws['B2'].alignment = Alignment(horizontal='left', vertical='top')

    ws['B3'].font = data_heading2
    ws['B3'].alignment = Alignment(horizontal='left', vertical='top')

    ws['B4'].font = data_heading2
    ws['B4'].alignment = Alignment(horizontal='left', vertical='top')

    ws['B6'].font = data_heading4
    ws['B6'].alignment = Alignment(horizontal='left', vertical='bottom')

    ws['F6'].font = data_heading4
    ws['F6'].alignment = Alignment(horizontal='left', vertical='bottom')

    ws['F7'].font = data_heading3
    ws['F7'].alignment = Alignment(horizontal='left', vertical='bottom')

    for i in range(7, 12):
        search_string = "B" + str(i)
        ws[search_string].font = data_heading2
        ws[search_string].alignment = Alignment(vertical='bottom')

    #formatting of heading row in table 1
    for i in range(66, 72):
        search_string2 = chr(i) + "12"
        ws[search_string2].font = table_headings
        ws[search_string2].alignment = Alignment(horizontal='center', vertical='center')
        ws[search_string2].fill = PatternFill("solid", fgColor="365194")
        ws[search_string2].border = Border(left=table_border, right=table_border, bottom=thin)

    #text formatting of table 1
    for i in range(13, 23):
        for j in range(66, 72):
            search_string = chr(j) + str(i)
            ws[search_string].font = data_heading2
            ws[search_string].alignment = Alignment(vertical='center', horizontal='center')

    #text formatting of below section of table 1
    for i in range(24, 28):
        search_string6 = "B" + str(i)
        ws[search_string6].font = data_heading2
        ws[search_string6].alignment = Alignment(vertical='center', horizontal='left')

        search_string3 = "F" + str(i)
        ws[search_string3].font = billing_data_headings
        ws[search_string3].alignment = Alignment(vertical='center', horizontal='center')

    """ws['G24'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G25'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G26'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G27'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G28'].border = Border(left=thin, right=thin)
    ws['G28'].fill = PatternFill("solid", fgColor="365194")"""

    for i in range(24, 28):
        search_string4 = "G" + str(i)
        ws[search_string4].font = billing_data
        ws[search_string4].alignment = Alignment(vertical='center', horizontal='center')

    ws['G28'].font = Font(name='Franklin Gothic Book',
                          size=11,
                          color="FFFFFF")
    ws['G28'].alignment = Alignment(vertical='center', horizontal='center')

    ws['E28'].font = Font(name='Franklin Gothic Book',
                          size=18,
                          color="365194")
    ws['E28'].alignment = Alignment(vertical='center', horizontal='center')

    for i in range(29, 33):
        search_string = "B" + str(i)
        ws[search_string].font = data_heading3
        ws[search_string].alignment = Alignment(vertical='bottom')

    for i in range(66, 72):
        for j in range(10):
            search_string5 = chr(i) + str(13 + j)
            if j == (len(invoice_details['Items']) - 1):
                ws[search_string5].border = Border(left=table_border, right=table_border, bottom=thick)
                if ((12 + j) % 2) == 1:
                    ws[search_string5].fill = PatternFill("solid", fgColor="F5F5F5")
                else:
                    pass
            else:
                ws[search_string5].border = Border(left=table_border, right=table_border, bottom=table_border)
                if ((12 + j) % 2) == 1:
                    ws[search_string5].fill = PatternFill("solid", fgColor="F5F5F5")
                else:
                    pass

    for i in range(66, 72):
        search_string7 = chr(i) + "22"
        ws[search_string7].border = Border(left=table_border, right=table_border, bottom=thick)

    # ==========================================================page #2=======================================================

    ws['B34'].font = data_heading1
    ws['B34'].alignment = Alignment(horizontal='left', vertical='bottom')

    ws['B35'].font = data_heading2
    ws['B35'].alignment = Alignment(horizontal='left', vertical='top')

    ws['B36'].font = data_heading2
    ws['B36'].alignment = Alignment(horizontal='left', vertical='top')

    ws['B37'].font = data_heading2
    ws['B37'].alignment = Alignment(horizontal='left', vertical='top')

    ws['B39'].font = data_heading4
    ws['B39'].alignment = Alignment(horizontal='left', vertical='bottom')

    ws['F39'].font = data_heading4
    ws['F39'].alignment = Alignment(horizontal='left', vertical='bottom')

    ws['F40'].font = data_heading3
    ws['F40'].alignment = Alignment(horizontal='left', vertical='bottom')

    for i in range(40, 45):
        search_string = "B" + str(i)
        ws[search_string].font = data_heading2
        ws[search_string].alignment = Alignment(vertical='bottom')

    for i in range(66, 72):
        search_string2 = chr(i) + "45"
        ws[search_string2].font = table_headings
        ws[search_string2].alignment = Alignment(horizontal='center', vertical='center')
        ws[search_string2].fill = PatternFill("solid", fgColor="365194")
        ws[search_string2].border = Border(left=table_border, right=table_border, bottom=thin)

    for i in range(46, 56):
        for j in range(66, 72):
            search_string = chr(j) + str(i)
            ws[search_string].font = data_heading2
            ws[search_string].alignment = Alignment(vertical='center', horizontal='center')

    for i in range(57, 61):
        search_string6 = "B" + str(i)
        ws[search_string6].font = data_heading2
        ws[search_string6].alignment = Alignment(vertical='center', horizontal='left')

        search_string3 = "F" + str(i)
        ws[search_string3].font = billing_data_headings
        ws[search_string3].alignment = Alignment(vertical='center', horizontal='center')

    ws['G57'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G58'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G59'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G60'].border = Border(left=thin, right=thin, bottom=table_border)
    ws['G61'].border = Border(left=thin, right=thin)
    ws['G61'].fill = PatternFill("solid", fgColor="365194")

    for i in range(57, 61):
        search_string4 = "G" + str(i)
        ws[search_string4].font = billing_data
        ws[search_string4].alignment = Alignment(vertical='center', horizontal='center')

    ws['G61'].font = Font(name='Franklin Gothic Book',
                          size=11,
                          color="FFFFFF")
    ws['G61'].alignment = Alignment(vertical='center', horizontal='center')

    ws['E61'].font = Font(name='Franklin Gothic Book',
                          size=18,
                          color="365194")
    ws['E61'].alignment = Alignment(vertical='center', horizontal='center')

    for i in range(62, 66):
        search_string = "B" + str(i)
        ws[search_string].font = data_heading3
        ws[search_string].alignment = Alignment(vertical='bottom')

    for i in range(66, 72):
        for j in range(len(invoice_details['Items'])-10):
            search_string5 = chr(i) + str(46 + j)
            if j == (len(invoice_details['Items']) - 1):
                ws[search_string5].border = Border(left=table_border, right=table_border, bottom=thick)
                if ((46 + j) % 2) == 1:
                    ws[search_string5].fill = PatternFill("solid", fgColor="F5F5F5")
                else:
                    pass
            else:
                ws[search_string5].border = Border(left=table_border, right=table_border, bottom=table_border)
                if ((46 + j) % 2) == 1:
                    ws[search_string5].fill = PatternFill("solid", fgColor="F5F5F5")
                else:
                    pass

    for i in range(66, 72):
        row = 46 + (len(invoice_details['Items'])-11)
        search_string8 = chr(i) + str(row)
        ws[search_string8].border = Border(left=table_border, right=table_border, bottom=thick)

    # =========================================================values used in invoice========================================

    invoice_number = "Invoice #" + str(invoice_details['Invoice_number'])
    date = "Date: " + invoice_details['Date']
    cgst = invoice_details['CGST'] + "%"
    sgst = invoice_details['SGST'] + "%"
    igst = invoice_details['IGST'] + "%"

    total = 0
    for i in invoice_details['Items']:
        product = int(i['Rate']) * int(i['Qty'])
        total += product

    cgst_cal = (int(invoice_details['CGST']) / 100) * total
    sgst_cal = (int(invoice_details['SGST']) / 100) * total
    igst_cal = (int(invoice_details['IGST']) / 100) * total
    final = total + cgst_cal + sgst_cal + igst_cal

    # =======================================================Values in invoice===============================================

    ws['B1'] = "Firm Name"
    ws['B2'] = "Firm Address"
    ws['B3'] = "Contact Numbers"
    ws['B4'] = "GST No. XXXXX-XXXXX"

    ws['B6'] = "Bill To"
    ws['F6'] = invoice_number
    ws['F7'] = date
    ws['B7'] = str(customer_details['Name']) + " | " + str(customer_details['Phone'])
    ws['B8'] = "State: " + str(customer_details['Address'])
    ws['B9'] = "StateCode: " + str(customer_details['Code'])
    ws['B10'] = "GSTN: " + str(customer_details['GSTN'])

    ws['B12'] = "S No."
    ws['C12'] = "HSN Code"
    ws['D12'] = "Description"
    ws['E12'] = "Qty."
    ws['F12'] = "Rate"
    ws['G12'] = "Amount"

    ws['B24'] = "Bank Name: Bank Name"
    ws['B25'] = "Account No.: XXXXXXXXXX"
    ws['B26'] = "Account Name: Account Name"
    ws['B27'] = "IFSC: IFSC Code"

    """ws['F24'] = "Subtotal"
    ws['G24'] = "=SUM(G10:G19)"
    ws['F25'] = "CGST(" + cgst + ")"
    ws['G25'] = cgst_cal
    ws['F26'] = "SGST(" + sgst + ")"
    ws['G26'] = sgst_cal
    ws['F27'] = "IGST(" + igst + ")"
    ws['G27'] = igst_cal
    ws['E28'] = "Total Cost"
    ws['G28'] = final"""

    ws['B29'] = "Make all checks payable to Company Name"
    ws['B30'] = "If you have any questions concerning this invoice, use the following contact information:"
    ws['B31'] = "Contact Details"
    ws['B32'] = "Thank you for your business!"

    ws['B34'] = "Firm Name"
    ws['B35'] = "Firm Details"
    ws['B36'] = "Contact Numbers"
    ws['B37'] = "GST No. XXXXX-XXXXX"

    ws['B39'] = "Bill To"
    ws['F39'] = invoice_number
    ws['F40'] = date
    ws['B40'] = str(customer_details['Name']) + " | " + str(customer_details['Phone'])
    ws['B41'] = "State: " + str(customer_details['Address'])
    ws['B42'] = "StateCode: " + str(customer_details['Code'])
    ws['B43'] = "GSTN: " + str(customer_details['GSTN'])

    ws['B45'] = "S No."
    ws['C45'] = "HSN Code"
    ws['D45'] = "Description"
    ws['E45'] = "Qty."
    ws['F45'] = "Rate"
    ws['G45'] = "Amount"

    ws['B57'] = "Bank Name: Bank Name"
    ws['B58'] = "Account No.: XXXXXXXXXX "
    ws['B59'] = "Account Name: Account Name"
    ws['B60'] = "IFSC: IFSC Code"

    ws['F57'] = "Subtotal"
    ws['G57'] = total
    ws['F58'] = "CGST(" + cgst + ")"
    ws['G58'] = cgst_cal
    ws['F59'] = "SGST(" + sgst + ")"
    ws['G59'] = sgst_cal
    ws['F60'] = "IGST(" + igst + ")"
    ws['G60'] = igst_cal
    ws['E61'] = "Total Cost"
    ws['G61'] = final

    ws['B62'] = "Make all checks payable to Company Name"
    ws['B63'] = "If you have any questions concerning this invoice, use the following contact information:"
    ws['B64'] = "Contact Details"
    ws['B65'] = "Thank you for your business!"

    # ======================================================Item Details in Table============================================
    start_row = 13
    start_col = 1
    serial_no = 1
    start_character = 69

    for i in range(0, 10):
        product_string = "=PRODUCT(" + chr(start_character) + str(start_row + i) + ", " + chr(
            start_character + 1) + str(start_row + i) + ")"
        ws.cell(row=(start_row + i), column=start_col + 1).value = serial_no
        serial_no += 1
        ws.cell(row=start_row + i, column=start_col + 2).value = int(invoice_details['Items'][i]['HSNCode'])
        ws.cell(row=start_row + i, column=start_col + 3).value = invoice_details['Items'][i]['ItemName']
        ws.cell(row=start_row + i, column=start_col + 4).value = int(invoice_details['Items'][i]['Qty'])
        ws.cell(row=start_row + i, column=start_col + 5).value = int(invoice_details['Items'][i]['Rate'])
        ws.cell(row=start_row + i, column=start_col + 6).value = product_string

    start_row2 = 46
    serial_no2 = 11

    for i in range(0, len(invoice_details['Items'])-10):
        product_string = "=PRODUCT(" + chr(start_character) + str(start_row2 + i) + ", " + chr(
            start_character + 1) + str(start_row2 + i) + ")"
        ws.cell(row=(start_row2 + i), column=start_col + 1).value = serial_no2
        serial_no2 += 1
        ws.cell(row=start_row2 + i, column=start_col + 2).value = int(invoice_details['Items'][10+i]['HSNCode'])
        ws.cell(row=start_row2 + i, column=start_col + 3).value = invoice_details['Items'][10+i]['ItemName']
        ws.cell(row=start_row2 + i, column=start_col + 4).value = int(invoice_details['Items'][10+i]['Qty'])
        ws.cell(row=start_row2 + i, column=start_col + 5).value = int(invoice_details['Items'][10+i]['Rate'])
        ws.cell(row=start_row2 + i, column=start_col + 6).value = product_string

    # ===============================================merging cells===========================================================

    ws.merge_cells('B1:E1')
    ws.merge_cells('B2:E2')
    ws.merge_cells('B3:E3')
    ws.merge_cells('B4:E4')

    ws.merge_cells('B6:E6')
    ws.merge_cells('F6:G6')
    ws.merge_cells('B7:E7')
    ws.merge_cells('F7:G7')
    ws.merge_cells('B8:D8')
    ws.merge_cells('B9:E9')
    ws.merge_cells('B10:E10')

    for i in range(24, 28):
        merge_range = "B" + str(i) + ":D" + str(i)
        ws.merge_cells(merge_range)

    ws.merge_cells('E28:F28')

    for i in range(29, 33):
        merge_range = "B" + str(i) + ":H" + str(i)
        ws.merge_cells(merge_range)

    ws.merge_cells('B34:E34')
    ws.merge_cells('B35:E35')
    ws.merge_cells('B36:E36')
    ws.merge_cells('B37:E37')

    ws.merge_cells('B39:E39')
    ws.merge_cells('F39:G39')
    ws.merge_cells('B40:E40')
    ws.merge_cells('F40:G40')
    ws.merge_cells('B41:D41')
    ws.merge_cells('B42:E42')
    ws.merge_cells('B43:E43')

    for i in range(57, 61):
        merge_range = "B" + str(i) + ":D" + str(i)
        ws.merge_cells(merge_range)

    ws.merge_cells('E61:F61')

    for i in range(62, 66):
        merge_range = "B" + str(i) + ":H" + str(i)
        ws.merge_cells(merge_range)

    # ====================================================saving file========================================================
    save_name = invoice_number + ".xlsx"
    wb.save(save_name)
    os.startfile(save_name, 'print')